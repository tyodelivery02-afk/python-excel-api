import json
import openpyxl
import tempfile
import os
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

# 检查某行是否在某个合并区域内，并返回该区域所有行
def get_merged_rows(ws, row_idx):
    merged_rows = set()
    for merged_range in ws.merged_cells.ranges:
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        if min_row <= row_idx <= max_row:
            for r in range(min_row, max_row + 1):
                merged_rows.add(r)
    return merged_rows

def process_excel(input_bytes, stats_json):
    """
    input_bytes: 上传的 Excel 文件 bytes
    stats_json: 前端传来的统计数据（JSON 字符串）
    返回: 处理后的 Excel bytes
    """
    # ---------------- 写入临时文件 ----------------
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(input_bytes)
        input_path = tmp.name

    try:
        # 解析 statsData
        stats_data = json.loads(stats_json)
    except Exception as e:
        os.remove(input_path)
        raise ValueError(f"统计数据解析失败: {e}")

    # 打开 Excel
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # 样式定义
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    bold_font = Font(name="Arial", size=16, bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    fixed_col_width = 10  # 列宽

    def style_cell(cell, val):
        cell.fill = yellow_fill
        cell.alignment = center_align
        cell.font = bold_font
        cell.border = thin_border
        if isinstance(val, (int, float)):
            cell.number_format = '0'
        else:
            cell.number_format = '@'
        cell.value = val
        # 设置列宽
        ws.column_dimensions[cell.column_letter].width = fixed_col_width

    # ---------------- 3. 新增列 & 写标题 ----------------
    start_col = 15  # O列
    stats_cols = stats_data[0]  # 标题行
    insert_count = len(stats_cols)  # 原来的列数

    ws.insert_cols(start_col, amount=insert_count)  # 一次插入所有列

    # 写标题（第4行）
    style_cell(ws.cell(row=4, column=start_col), "全体")
    for i, header in enumerate(stats_cols[1:]):  # 从第2列开始
        style_cell(ws.cell(row=4, column=start_col + 1 + i), header)

    # ---------------- 4. 按マスタ番号匹配写入数据 & 隐藏行 ----------------
    stats_rows = stats_data[1:-2]
    total_rows = stats_data[-2:]

    for row_idx in range(5, ws.max_row + 1):
        master_no = ws.cell(row=row_idx, column=7).value  # G列
        match = next((r for r in stats_rows if r[0] == master_no), None)

        if match:
            val2 = float(match[1]) if match[1] not in [None, ""] else 0
            val3 = float(match[2]) if match[2] not in [None, ""] else 0
            total_val = val2 + val3
            style_cell(ws.cell(row_idx, column=start_col), total_val)
            for i, val in enumerate(match[1:]):
                style_cell(ws.cell(row_idx, column=start_col + 1 + i), val)
            ws.row_dimensions[row_idx].hidden = False
        else:
            row_cells = ws[row_idx]
            is_empty = all(cell.value in [None, ""] for cell in row_cells)

            # 找到该行属于哪些合并区域
            merged_rows = get_merged_rows(ws, row_idx)

            if is_empty:
                # 显示所有受影响行
                for r in (merged_rows or {row_idx}):
                    ws.row_dimensions[r].hidden = False
            else:
                # 隐藏所有受影响行
                for r in (merged_rows or {row_idx}):
                    ws.row_dimensions[r].hidden = True

    # ---------------- 5. 写入总计行 ----------------
    def is_row_empty(row):
        return all(cell.value in [None, ""] for cell in row)

    last_data_row = 4
    for row_idx in range(5, ws.max_row + 1):
        row = ws[row_idx]
        if not is_row_empty(row):
            last_data_row = row_idx

    for idx, t_row in enumerate(total_rows):
        new_row_idx = last_data_row + 1 + idx
        ws.cell(row=new_row_idx, column=14, value=t_row[0])
        if idx == 0:
            total_val = float(total_rows[1][1] or 0) + float(total_rows[1][2] or 0)
        else:
            val2 = float(t_row[1] or 0)
            val3 = float(t_row[2] or 0)
            total_val = val2 + val3
        style_cell(ws.cell(new_row_idx, column=start_col), total_val)
        for j, val in enumerate(t_row[1:]):
            style_cell(ws.cell(new_row_idx, column=start_col + 1 + j), val)

    # ---------------- 6. 代理店计算（竖向写入） ----------------
    agent_map = {}
    for row_idx in range(5, ws.max_row + 1):
        if ws.row_dimensions[row_idx].hidden:
            continue
        agent = ws.cell(row=row_idx, column=8).value
        tokyo_val = ws.cell(row=row_idx, column=start_col + 2).value
        try:
            tokyo_val = float(tokyo_val or 0)
        except:
            tokyo_val = 0
        agent_map[agent] = agent_map.get(agent, 0) + tokyo_val

    agent_div = {"CAINIAO-E": 450, "TEMU": 250, "MMA-CN": 180}
    insert_row_num = last_data_row + 1 + len(total_rows)

    cell_title = ws.cell(row=insert_row_num, column=14, value="代理店")
    cell_title.fill = orange_fill
    cell_title.alignment = center_align
    cell_title.font = bold_font
    cell_title.border = thin_border

    cell_title = ws.cell(row=insert_row_num, column=17, value="かご数")
    cell_title.fill = orange_fill
    cell_title.alignment = center_align
    cell_title.font = bold_font
    cell_title.border = thin_border

    start_row = insert_row_num + 1
    for idx, (agent, total) in enumerate(agent_map.items()):
        div = agent_div.get(agent)
        if div:
            val = (total // div) + (1 if total % div else 0)
        else:
            val = "計算外"

        cell_agent = ws.cell(row=start_row + idx, column=14, value=agent)
        cell_agent.fill = orange_fill
        cell_agent.alignment = center_align
        cell_agent.font = bold_font
        cell_agent.border = thin_border

        cell_val = ws.cell(row=start_row + idx, column=17, value=val)
        cell_val.fill = orange_fill
        cell_val.alignment = center_align
        cell_val.font = bold_font
        cell_val.border = thin_border

    # ---------------- 7. 保存并返回 bytes ----------------
    output_path = input_path.replace(".xlsx", "_filled.xlsx")
    wb.save(output_path)

    with open(output_path, "rb") as f:
        output_bytes = f.read()

    os.remove(input_path)
    os.remove(output_path)

    return output_bytes
